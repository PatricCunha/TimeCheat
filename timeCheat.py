from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Style
from openpyxl.compat import range
#import datetime
from datetime import date
from datetime import time
import calendar
import sys

def name_to_num(abbr):
	return list(calendar.month_name).index(abbr)
def num_to_name(num):
	return calendar.month_name[num]
def diff_month_date(arr, month, start):
	lastday = calendar.monthrange(date.today().year,month)[1]
	otherMonthWeeks = currentCalendar.monthdays2calendar(date.today().year, month)
	if(start):
		otherMonthWeekDays = otherMonthWeeks[-1]
	else:
		otherMonthWeekDays = otherMonthWeeks[0]
	for day in otherMonthWeekDays:
		print day
		print arr
		if(day[1] == arr[1]):
			return day[0]
# check if there is a filename given
assert len(sys.argv) > 1, "Filename not included in arguments"
# get the name from the first argument
filename = sys.argv[1]
# load the workbook
wb = load_workbook(filename)
# get existing sheets
sheets = wb.get_sheet_names()
# last month
lastmonth = sheets[-1]
# as a numeric value
lastmonthnum = name_to_num(lastmonth)
# get this month as a num
thismonthnum = lastmonthnum + 1
# get the next month as num
nextmonthnum = thismonthnum + 1
# make sure we reset on december
if lastmonthnum > 11 :
	thismonthnum = 1
# get the month from the number
month = num_to_name(thismonthnum)
print month
# create a new worksheet
ws = wb.create_sheet()
# change the title to the month
ws.title = month
# print month to console
calendar.prmonth(2015, thismonthnum)
# new calendar starting each week on monday
currentCalendar = calendar.Calendar(0)
# let's create headers
ws["A1"] = "Day"
ws["B1"] = "Start"
ws["C1"] = "End"
ws["D1"] = "Total"
ws["E1"] = "Difference"
ws["H1"] = "Workday"
ws["H2"] = time(9,0,0,0)
ws["H2"].number_format = 'hh:mm:ss'
ws["I1"] = "Weekend"
ws["I2"] = time(0,0,0,0)
ws["I2"].number_format = 'hh:mm:ss'
lineCounter = 2
workStyleFill = PatternFill(patternType='solid',fgColor='C6EFCE')
workStyleFont = Font(name='Calibri',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='006100')
workStyle = Style(fill=workStyleFill, font=workStyleFont)
noWorkStyleFill = PatternFill(patternType='solid',fgColor='FFEB9C')
noWorkStyleFont = Font(name='Calibri',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='9C6500')
noWorkStyle = Style(fill=noWorkStyleFill, font=noWorkStyleFont)
# get datetime iterator for each day in month
monthdates = currentCalendar.itermonthdays2(date.today().year, thismonthnum)
monthStr = '='
trueDate = 0
for day in monthdates:
	countStr = str(lineCounter)
	if(day[0] == 0):
		if(lineCounter <= 9):
			trueDate = diff_month_date(day, lastmonthnum, True)
		else:
			trueDate = diff_month_date(day, nextmonthnum, False)
	else:
		trueDate = day[0]
	if(day[1] == 5 or day[1] == 6):
		cell = "A" + countStr
		ws[cell] = trueDate
		ws[cell].style = noWorkStyle
		cell = "B" + countStr
		ws[cell].style = noWorkStyle
		cell = "C" + countStr
		ws[cell].style = noWorkStyle
		cell = "D" + countStr
		ws[cell].style = noWorkStyle
		cell = "E" + countStr
		ws[cell].style = noWorkStyle
	else:
		cell = "A" + countStr
		ws[cell] = trueDate
		cell = "B" + countStr
		ws[cell] = time(9,0,0,0)
		cell = "C" + countStr
		ws[cell] = time(18,0,0,0)
		cell = "D" + countStr
		ws[cell] = "=C"+countStr +"-B"+countStr
		ws[cell].number_format = 'hh:mm:ss'
		cell = "E" + countStr
		ws[cell] = "=D"+countStr+"-H2"
		ws[cell].number_format = 'hh:mm:ss'
	if day[1] == 4:
		lineCounter += 1
		countStr = str(lineCounter)
		cell = "A" + countStr
		ws[cell] = "Week"
		cell = "E" + countStr
		monthStr +=cell+'+'
		ws[cell] = "=SUM(E"+str(lineCounter-5)+":E"+str(lineCounter-1)+")"
		ws[cell].number_format = 'hh:mm:ss'
	lineCounter += 1
	print day
countStr = str(lineCounter)
cell = "A" + countStr
ws[cell] = month
cell = "E" + countStr
#ws[cell] = "=E7+E15+E23+E31+E39"
ws[cell] = monthStr[:-1]
ws[cell].number_format = 'hh:mm:ss'

# overwrites same file, filename is required
wb.save(filename)